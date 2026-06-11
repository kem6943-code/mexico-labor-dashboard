# -*- coding: utf-8 -*-
"""
parsers.py — 멕시코 인건비 데이터 파싱 모듈
==============================================
CONTPAQi Nóminas 엑셀 파일을 pandas DataFrame으로 변환합니다.

왜 이렇게 복잡한가?
  → CONTPAQi 급여대장은 "보고서" 형태라서, 일반 테이블이 아닙니다.
    헤더가 8행에 있고, 부서 구분행·소계행이 데이터 사이에 끼어 있습니다.
    이 파서는 그 구조를 자동 인식하여 깔끔한 테이블로 추출합니다.
"""

import pandas as pd
import openpyxl
import os
import re
from typing import Dict, List, Tuple, Optional


# ===================================================================
# 한국어 컬럼 매핑 사전
# ===================================================================

# 급여 지급항목 (Percepciones)
EARNING_MAP = {
    'Sueldo': '기본급',
    'Séptimo día': '주휴수당',
    'Horas extras': '초과근무수당',
    'Destajos': '성과급',
    'Prima dominical': '일요근로수당',
    'Día festivo / descanso': '휴일근로수당',
    'Gratificación': '보너스',
    'Compensación': '보상수당',
    'Premios eficiencia': '능률수당',
    'Bono puntualidad': '출근수당',
    'Vacaciones a tiempo': '연차수당',
    'Prima de vacaciones a tiempo': '연차보너스',
    'Vacaciones reportadas $': '연차정산금',
    'Prima de vacaciones reportada $': '연차정산보너스',
    'Aguinaldo': '크리스마스보너스',
    'Fondo ahorro empresa': '회사저축기금',
    'Despensa': '식대',
    'Fondo de Ahorro Empleado': '직원저축기금',
    'Complemento': '보충수당',
    'Premio de asistencia mensual': '월출근수당',
}

# 공제항목 (Deducciones)
DEDUCTION_MAP = {
    'Ret. Inv. Y Vida': '퇴직/사망보험',
    'Ret. Cesantia': '실업보험',
    'Ret. Enf. y Mat. obrero': '의료보험(직원)',
    'Seguro de vivienda Infonavit': '주택보험',
    'Préstamo infonavit (FD)': '주택대출(FD)',
    'Préstamo infonavit (CF)': '주택대출(CF)',
    'Préstamo infonavit (PORC)': '주택대출(%)',
    'Subs al Empleo acreditado': '고용보조금(적용)',
    'Subs al Empleo (mes)': '고용보조금(월)',
    'I.S.R. antes de Subs al Empleo': 'ISR(보조금전)',
    'I.S.R. Art174': 'ISR(Art174)',
    'I.S.R. (mes)': '소득세(ISR)',
    'I.M.S.S.': '사회보험(IMSS)',
    'Cuota sindical': '조합비',
    'Préstamo FONACOT': 'FONACOT대출',
    'Aportacion F.A. Empledo': '직원저축공제',
    'Aportacion F.A. Empresa': '회사저축공제',
    'Ajuste en Subsidio para el empleo': '고용보조금조정',
    'Subs entregado que no correspondía': '부당보조금환수',
    'Ajuste al neto': '실수령조정',
    'ISR de ajuste mensual': 'ISR월말정산',
    'ISR ajustado por subsidio': 'ISR보조금정산',
    'Ajuste al Subsidio Causado': '보조금원인정산',
    'Pension Alimenticia': '양육비',
    'Gastos Medicos': '의료비',
    'Infonavit CF correspondiente': 'Infonavit CF',
    'Infonavit FD correspondiente': 'Infonavit FD',
}

# 회사 의무부담금 (Obligaciones Patronales)
OBLIGATION_MAP = {
    'Invalidez y Vida': '장해/사망(회사)',
    'Cesantia y Vejez': '실업/연금(회사)',
    'Enf. y Mat. Patron': '의료(회사)',
    '2% Fondo retiro SAR (8)': '퇴직기금SAR(2%)',
    '3% Impuesto estatal': '지방세(3%)',
    'Riesgo de trabajo (9)': '산재보험',
    'I.M.S.S. empresa': 'IMSS(회사)',
    'Infonavit empresa': 'Infonavit(회사)',
    'Guarderia I.M.S.S. (7)': '보육비(1%)',
}

# 부서명 한국어 매핑
DEPT_MAP = {
    'INYECCION': '사출',
    'INYECCION II': '사출2',
    'ENSAMBLE': '조립',
    'CALIDAD': '품질',
    'EMBARQUES': '출하',
    'RECURSOS HUMANOS': '인사',
    'RECURSOS HUMANO': '인사',
    'MANTENIMIENTO INYECCION': '사출유지보수',
    'MANTENIMIENTO MOLDES': '금형유지보수',
    'MANTENIMIENTO GRAL': '일반유지보수',
    'ADMINISTRACION': '관리',
    'CONTABILIDAD': '회계',
    'CONTABILIDAD II': '회계2',
    'ALMACEN Y MATERIALES': '자재/창고',
    'ALMACEN': '창고',
    'INNOVACION Y DESARROLLO': '혁신개발',
    'INNOVACION Y DESARROLL': '혁신개발',
    'PRODUCTION CONTROL': '생산관리',
    'SISTEMAS': '시스템',
    'DIRECCION': '경영',
    'PLANTA 3': '제3공장',
    'PLANTA 3 DIRECCION': '제3공장경영',
    'PLANTA 3 INYECCION': '제3공장사출',
}

# 직위명 한국어 매핑
POSITION_MAP = {
    'AYUDANTE GENERAL': '일반보조',
    'OPERARIO': '오퍼레이터',
    'INSPECTOR CALIDAD': '품질검사원',
    'RESIDENTE': '주재원',
    'TECNICO EN MTTOX': '유지보수기술자',
    'AVANCE': '진행관리',
    'LIDER DE ENSAMBLE': '조립리더',
    'CAPTURISTA': '데이터입력',
    'MONTADOR DE MOLDES': '금형장착기사',
    'LIMPIEZA': '청소',
    'TECNICO': '기술자',
    'LIDER DE INYECCION': '사출리더',
    'SUPERVISOR DE PRODUCCION': '생산감독',
    'LIDER DE CALIDAD': '품질리더',
    'LIDER DE MATERIALES': '자재리더',
    'PLANNER': '플래너',
    'SUPERVISOR DE EMBARQUES': '출하감독',
    'PRESIDENTE': '대표',
    'DIRECTOR DE CALIDAD': '품질이사',
    'SUB LIDER': '부리더',
    'DIRECTOR GENERAL': '사장',
    'GERENTE DE PRODUCCION': '생산부장',
    'JEFE DE RECURSOS HUMANOS': '인사팀장',
    'CONTADOR GENERAL': '회계팀장',
}


def translate_dept(dept_name: str) -> str:
    """부서명을 한국어로 변환. 매핑 없으면 원문 유지."""
    if not dept_name:
        return '미분류'
    dept_upper = dept_name.strip().upper()
    return DEPT_MAP.get(dept_upper, dept_name)


def translate_position(pos_name: str) -> str:
    """직위명을 한국어로 변환."""
    if not pos_name:
        return '미분류'
    pos_upper = pos_name.strip().upper()
    return POSITION_MAP.get(pos_upper, pos_name)


# ===================================================================
# 1. 직원 카탈로그 파서
# ===================================================================

def parse_employee_catalog(filepath: str) -> pd.DataFrame:
    """
    Catalogo de Empleados.xlsx를 파싱하여 활성 직원 DataFrame을 반환합니다.
    
    왜 ACTIVO 필터링을 하는가?
      → 카탈로그에는 퇴사자(BAJA) 9,969명이 포함되어 있어서 
        현재 인원 분석에는 활성(ACTIVO) 직원만 필요합니다.
    
    성능 최적화:
      → pd.read_excel 사용 (openpyxl 셀 단위 읽기 대비 10배 이상 빠름)
    """
    # pandas read_excel로 한번에 읽기 (cell-by-cell 대비 10배+ 빠름)
    df = pd.read_excel(filepath, sheet_name='Empleados', engine='openpyxl')
    
    # 최근 데이터(2025년 이후 퇴사 등) 포함하거나 조건없이 전체 반환 (메모리 문제 없으므로)
    if 'Fecha de baja' in df.columns:
        df['Fecha_de_baja_dt'] = pd.to_datetime(df['Fecha de baja'], errors='coerce')
        if 'Estatus empleado' in df.columns:
            df = df[
                df['Estatus empleado'].str.upper().str.contains('ACTIV', na=False) |
                (df['Fecha_de_baja_dt'] >= '2025-01-01')
            ].copy()
        df = df.drop(columns=['Fecha_de_baja_dt'], errors='ignore')
    
    # 핵심 컬럼 한국어 rename
    rename_map = {
        'Código': '사원번호',
        'Nombre': '이름',
        'Apellido paterno': '성(부)',
        'Apellido materno': '성(모)',
        'Departamento': '부서(원문)',
        'Puesto': '직위(원문)',
        'Salario diario': '일급',
        'Tipo de periodo': '급여주기',
        'Turno de trabajo': '근무교대',
        'Sindicalizado': '노조가입',
        'Fecha de alta': '입사일',
        'Fecha de baja': '퇴사일',
        'Sexo': '성별',
        'Estatus empleado': '재직상태',
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
    
    # 한국어 부서/직위 컬럼 추가
    if '부서(원문)' in df.columns:
        df['부서'] = df['부서(원문)'].apply(translate_dept)
    if '직위(원문)' in df.columns:
        df['직위'] = df['직위(원문)'].apply(translate_position)
    
    # 풀네임 생성
    name_cols = ['성(부)', '성(모)', '이름']
    existing_name_cols = [c for c in name_cols if c in df.columns]
    if existing_name_cols:
        df['직원명'] = df[existing_name_cols].fillna('').apply(
            lambda x: ' '.join(x).strip(), axis=1
        )
    
    # 급여주기 한국어화
    period_map = {'Semanal': '주급', 'Quincenal': '격주급', 'MENSUAL': '월급', 'Mensual': '월급'}
    if '급여주기' in df.columns:
        df['급여주기(한)'] = df['급여주기'].map(period_map).fillna(df['급여주기'])
    
    # 노조 한국어화
    if '노조가입' in df.columns:
        df['노조가입(한)'] = df['노조가입'].map({'S': '조합원', 'C': '비조합원'}).fillna('미분류')
    
    # 교대 한국어화
    shift_map = {'Mixto': '교대', 'Nocturno': '야간', 'Matutino': '주간', 'Vesperino': '오후'}
    if '근무교대' in df.columns:
        df['근무교대(한)'] = df['근무교대'].map(shift_map).fillna(df['근무교대'])
    
    df = df.reset_index(drop=True)
    return df


# ===================================================================
# 2. 급여대장 파서 (Lista de raya)
# ===================================================================

def _clean_header(val: str) -> str:
    """헤더 셀의 별표(*) 장식 제거."""
    if not val:
        return ''
    return re.sub(r'\*', '', str(val)).strip()


def _find_header_row(ws) -> int:
    """
    헤더 행 번호를 찾습니다.
    'Código'가 A열에 있는 행을 헤더로 인식합니다.
    """
    for r in range(1, 20):
        val = ws.cell(row=r, column=1).value
        if val and 'Código' in str(val):
            return r
    return 8  # 기본값 (대부분 8행)


def _extract_period_info(ws) -> Dict:
    """
    급여대장 상단에서 기간 정보를 추출합니다.
    예: "Periodo 1 al 1 Semanal del 29/12/2025 al 04/01/2026"
    """
    info = {
        'company': '',
        'period_type': '',
        'period_number': '',
        'date_from': '',
        'date_to': '',
        'reg_pat': '',
        'fte_value': 0.0,
    }
    
    for r in range(1, 10):
        for c in range(1, 10):
            val = ws.cell(row=r, column=c).value
            if not val:
                continue
            val_str = str(val)
            
            if 'DONGJIN' in val_str.upper():
                info['company'] = val_str
            
            if 'PERIODO' in val_str.upper():
                # 급여 주기 타입 추출
                if 'SEMANAL' in val_str.upper():
                    info['period_type'] = '주급'
                    info['fte_value'] = 0.25
                elif 'QUINCENAL' in val_str.upper():
                    info['period_type'] = '격주급'
                    info['fte_value'] = 0.5
                elif 'MENSUAL' in val_str.upper():
                    info['period_type'] = '월급'
                    info['fte_value'] = 1.0
                
                # 기간 번호 추출
                period_match = re.search(r'Periodo\s+(\d+)', val_str, re.IGNORECASE)
                if period_match:
                    info['period_number'] = period_match.group(1)
                
                # 날짜 추출
                date_match = re.search(r'del\s+(\S+)\s+al\s+(\S+)', val_str, re.IGNORECASE)
                if date_match:
                    info['date_from'] = date_match.group(1)
                    info['date_to'] = date_match.group(2)
            
            if 'Reg Pat' in val_str:
                info['reg_pat'] = val_str.replace('Reg Pat IMSS:', '').strip()
    
    return info


def parse_payroll_file(filepath: str) -> Tuple[pd.DataFrame, Dict]:
    """
    CONTPAQi 급여대장 엑셀 파일을 파싱합니다.
    
    Returns:
        (DataFrame, period_info_dict)
        
    왜 이렇게 파싱하는가?
      → CONTPAQi 보고서는 부서별로 그룹핑되어 있고,
        "Departamento XX 부서명" 행과 "Total Depto" 행이 데이터 사이에 있습니다.
        이를 제거하고 부서 컬럼을 추가하여 깔끔한 테이블로 만듭니다.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Hoja1']
    
    # 기간 정보 추출
    period_info = _extract_period_info(ws)
    
    # 헤더 행 찾기
    header_row = _find_header_row(ws)
    
    # 헤더 추출
    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        cleaned = _clean_header(val) if val else f'Col_{c}'
        headers.append(cleaned)
    
    # 데이터 추출: 부서 구분행을 인식하여 부서 컬럼 추가
    data = []
    current_dept = ''
    current_reg_pat = ''
    
    for r in range(header_row + 1, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value
        
        if not a_val and not b_val:
            continue
        
        a_str = str(a_val).strip() if a_val else ''
        
        # 레지스트리 패턴 행 (Reg. Pat. IMSS: ...)
        if 'Reg. Pat' in a_str:
            reg_match = re.search(r'D\d+', a_str)
            if reg_match:
                current_reg_pat = reg_match.group(0)
            continue
        
        # 부서 헤더 행 (Departamento XX 부서명)
        if a_str.startswith('Departamento'):
            # "Departamento 11 RECURSOS HUMANOS" → "RECURSOS HUMANOS"
            dept_match = re.match(r'Departamento\s+\d+\s+(.*)', a_str)
            if dept_match:
                current_dept = dept_match.group(1).strip()
            else:
                current_dept = a_str.replace('Departamento', '').strip()
            continue
        
        # 소계/합계 행 건너뛰기
        if 'Total' in a_str and ('Depto' in a_str or 'Gral' in a_str):
            continue
        
        # 구분선 행 건너뛰기
        if a_str.startswith('---') or a_str.startswith('==='):
            continue
        
        # 빈 코드 건너뛰기
        if not a_val or a_str == '':
            continue
            
        # 실제 직원 데이터 행인지 확인 (사원번호 패턴)
        if not re.match(r'^\d{3,5}$', a_str):
            continue
        
        # 직원 데이터 행 추출
        fte_val = period_info.get('fte_value', 0.0)
        row_data = {'부서(원문)': current_dept, 'Reg_Pat': current_reg_pat, 'FTE_기여도': fte_val}
        for c in range(1, ws.max_column + 1):
            col_name = headers[c-1]
            cell_val = ws.cell(row=r, column=c).value
            row_data[col_name] = cell_val
        data.append(row_data)
    
    wb.close()
    
    if not data:
        return pd.DataFrame(), period_info
    
    df = pd.DataFrame(data)
    
    # 한국어 부서 컬럼 추가
    df['부서'] = df['부서(원문)'].apply(translate_dept)
    
    # 핵심 컬럼 rename
    if 'Código' in df.columns:
        df = df.rename(columns={'Código': '사원번호'})
    if 'Empleado' in df.columns:
        df = df.rename(columns={'Empleado': '직원명'})
    
    # 숫자 컬럼 타입 변환 (합계 관련 컬럼들)
    numeric_cols = [c for c in df.columns if c not in [
        '사원번호', '직원명', '부서(원문)', '부서', 'Reg_Pat',
        'Col_1', 'Col_2'
    ]]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # 총 지급/공제/실수령 컬럼 식별 및 rename
    rename_summary = {}
    for col in df.columns:
        if 'TOTAL' in str(col).upper() and 'PERCEPCIONES' in str(col).upper():
            rename_summary[col] = '총지급액'
        elif 'TOTAL' in str(col).upper() and 'DEDUCCIONES' in str(col).upper():
            rename_summary[col] = '총공제액'
        elif 'NETO' in str(col).upper() and len(str(col)) < 10:
            rename_summary[col] = '실수령액'
        elif 'TOTAL' in str(col).upper() and 'OBLIGACIONES' in str(col).upper():
            rename_summary[col] = '총회사부담금'
        elif 'Otras' in str(col) and 'Percepciones' in str(col):
            rename_summary[col] = '기타수당'
        elif 'Otras' in str(col) and 'Deducciones' in str(col):
            rename_summary[col] = '기타공제'
        elif 'Otras' in str(col) and 'Obligaciones' in str(col):
            rename_summary[col] = '기타회사부담'
    
    df = df.rename(columns=rename_summary)
    
    # 지급/공제/회사부담 개별 항목 한국어화
    earning_rename = {k: v for k, v in EARNING_MAP.items() if k in df.columns}
    deduction_rename = {k: v for k, v in DEDUCTION_MAP.items() if k in df.columns}
    obligation_rename = {k: v for k, v in OBLIGATION_MAP.items() if k in df.columns}
    
    df = df.rename(columns={**earning_rename, **deduction_rename, **obligation_rename})
    
    return df, period_info


# ===================================================================
# 3. 월간 합산 로직
# ===================================================================

def load_all_payroll_files(data_dir: str) -> Dict[str, List[Tuple[pd.DataFrame, Dict]]]:
    """
    데이터 폴더에서 모든 급여대장을 로드하여 유형별로 그룹핑합니다.
    
    Returns:
        {'주급': [(df1, info1), (df2, info2), ...],
         '격주급': [...],
         '월급': [...]}
    """
    result = {'주급': [], '격주급': [], '월급': []}
    
    for fname in sorted(os.listdir(data_dir)):
        if not fname.endswith('.xlsx') or 'Catalogo' in fname:
            continue
        
        fpath = os.path.join(data_dir, fname)
        try:
            df, info = parse_payroll_file(fpath)
            if df.empty:
                continue
            
            info['filename'] = fname
            period_type = info.get('period_type', '')
            if period_type in result:
                result[period_type].append((df, info))
        except Exception as e:
            print(f"[경고] 파일 파싱 실패: {fname} → {e}")
    
    return result


def merge_weekly_to_monthly(payroll_data: Dict) -> pd.DataFrame:
    """
    주급(4주분) + 격주급(2기분) + 월급을 합산하여
    직원별 월간 총 인건비 DataFrame을 생성합니다.
    
    왜 단순 합산하는가?
      → 주급 직원은 한 달에 4번(주1~4), 격주급은 2번(Q1~2), 
        월급은 1번 지급됩니다. 각각을 합산하면 월 기준 비용이 됩니다.
    """
    all_frames = []
    
    for period_type, entries in payroll_data.items():
        for df, info in entries:
            df_copy = df.copy()
            df_copy['급여유형'] = period_type
            df_copy['기간번호'] = info.get('period_number', '')
            
            # 1. 파일명에서 YYYY_MM 추출 시도
            fname = info.get('filename', '')
            ym_match = re.search(r'(\d{4})_(\d{2})', fname)
            
            # 2. 파일명 추출 실패 시 엑셀 내부에 적힌 급여 종료일(date_to) 추출 (DD/MM/YYYY)
            date_to = str(info.get('date_to', '')).strip()
            date_to_clean = re.sub(r'[.-]', '/', date_to) # 통일된 구분자
            
            if ym_match:
                df_copy['YearMonth'] = f"{ym_match.group(1)}-{ym_match.group(2)}"
            elif date_to_clean and '/' in date_to_clean:
                parts = date_to_clean.split('/')
                if len(parts) >= 3:
                    year_match = re.search(r'(20\d\d)', parts[2])
                    year = year_match.group(1) if year_match else '2026'
                    month = ''.join(filter(str.isdigit, parts[1]))
                    month = month.zfill(2) if month else '01'
                    df_copy['YearMonth'] = f"{year}-{month}"
                else:
                    df_copy['YearMonth'] = '2026-01' # 강제 회복 fallback
            else:
                # 3. 최후의 수단: 파일명에 'Mensual 1 2026' 같은 텍스트가 있다면...
                # 못찾으면 그냥 2026-01으로 할당 (전개 편의성)
                y_match = re.search(r'(20\d\d)', fname)
                year = y_match.group(1) if y_match else '2026'
                df_copy['YearMonth'] = f"{year}-01"
                
            all_frames.append(df_copy)
    
    if not all_frames:
        return pd.DataFrame()
    
    # 모든 데이터 결합
    combined = pd.concat(all_frames, ignore_index=True, sort=False)
    combined = combined.fillna(0)
    
    # 문자열 컬럼 식별
    str_cols = ['사원번호', '직원명', '부서(원문)', '부서', 'Reg_Pat', '급여유형', '기간번호']
    str_cols = [c for c in str_cols if c in combined.columns]
    
    # 숫자 컬럼 식별
    numeric_cols = [c for c in combined.columns if c not in str_cols]
    for col in numeric_cols:
        combined[col] = pd.to_numeric(combined[col], errors='coerce').fillna(0)
    
    # NaN 결측치 처리 및 문자열 명시적 변환
    if 'YearMonth' not in combined.columns:
        combined['YearMonth'] = '2026-01'
    
    combined['YearMonth'] = combined['YearMonth'].fillna('2026-01').astype(str).str.strip()
    combined.loc[combined['YearMonth'].str.lower().isin(['unknown', 'nan', 'nat', '']), 'YearMonth'] = '2026-01'
    
    # pd.to_datetime()을 통해 유효하지 않은 포맷을 걸러내고 통일
    temp_dates = pd.to_datetime(combined['YearMonth'], errors='coerce')
    combined['YearMonth'] = temp_dates.dt.strftime('%Y-%m').fillna('2026-01')
        
    # 직원별 월간 합산
    # 그룹핑 키: 사원번호 + 부서 + YearMonth
    group_keys = ['사원번호', '직원명', '부서(원문)', '부서', 'YearMonth']
    group_keys = [k for k in group_keys if k in combined.columns]
    
    # 합산할 숫자 컬럼
    sum_cols = [c for c in numeric_cols if c not in ['기간번호']]
    
    monthly = combined.groupby(group_keys, as_index=False)[sum_cols].sum()
    
    # 급여유형 정보도 보존 (첫 번째 값)
    if '급여유형' in combined.columns:
        type_map = combined.groupby(group_keys, as_index=False)['급여유형'].first()
        # group_keys로 머지
        monthly = monthly.merge(type_map, on=group_keys, how='left')
    
    return monthly


def calculate_summary_stats(monthly_df: pd.DataFrame, employee_df: pd.DataFrame = None) -> Dict:
    """
    월간 합산 데이터에서 핵심 KPI를 계산합니다.
    
    Returns:
        {
            'total_headcount': 총 인원,
            'total_earnings': 총 지급액,
            'total_deductions': 총 공제액,
            'total_net': 총 실수령액,
            'total_obligations': 총 회사부담금,
            'total_labor_cost': 총 인건비 (지급 + 회사부담),
            'avg_cost_per_person': 인당 평균 비용,
            'dept_summary': 부서별 요약 DataFrame,
        }
    """
    stats = {}
    
    stats['total_headcount'] = len(monthly_df)
    
    # 정확한 입/퇴사일 기반 FTE 계산
    if employee_df is not None and not employee_df.empty and '사원번호' in employee_df.columns:
        emp_meta = employee_df[['사원번호', '입사일', '퇴사일']].copy() if '퇴사일' in employee_df.columns else employee_df[['사원번호', '입사일']].copy()
        if '퇴사일' not in emp_meta.columns:
            emp_meta['퇴사일'] = None
        if '입사일' not in emp_meta.columns:
            emp_meta['입사일'] = None
            
        emp_meta['사원번호'] = emp_meta['사원번호'].astype(str).str.strip()
        
        monthly_tmp = monthly_df.copy()
        monthly_tmp['사원번호'] = monthly_tmp['사원번호'].astype(str).str.strip()
        merged = monthly_tmp.merge(emp_meta, on='사원번호', how='left')
        
        # 1) 입사일/퇴사일을 파싱 (오류 시 NaT)
        start_dt = pd.to_datetime(merged['입사일'], errors='coerce')
        end_dt = pd.to_datetime(merged['퇴사일'], errors='coerce')
        
        # 2) 해당 월(YearMonth)의 시작일과 말일 계산
        ym_series = merged['YearMonth'].astype(str).replace('Unknown', '2026-01')
        m_start = pd.to_datetime(ym_series + '-01', errors='coerce').fillna(pd.to_datetime('2026-01-01'))
        m_end = m_start + pd.offsets.MonthEnd(0)
        
        # 3) 입사일/퇴사일 결측치 채우기 (해당 월의 처음/끝으로 간주)
        start_dt = start_dt.fillna(m_start)
        end_dt = end_dt.fillna(m_end)
        
        # 4) 벡터 연산으로 근무 시작일/종료일 클리핑 (max/min)
        actual_start = start_dt.clip(lower=m_start)
        actual_end = end_dt.clip(upper=m_end)
        
        # 5) 근무 일수 계산 및 음수 값은 0으로 처리
        month_days = m_end.dt.day
        worked_days = (actual_end - actual_start).dt.days + 1
        worked_days = worked_days.clip(lower=0)
        
        monthly_df['FTE'] = worked_days / month_days
    else:
        # fallback
        monthly_df['FTE'] = monthly_df['FTE_기여도'] if 'FTE_기여도' in monthly_df.columns else 1.0

    stats['total_fte'] = monthly_df['FTE'].sum()
    
    # 총 지급액
    if '총지급액' in monthly_df.columns:
        stats['total_earnings'] = monthly_df['총지급액'].sum()
    else:
        stats['total_earnings'] = 0
    
    # 총 공제액
    if '총공제액' in monthly_df.columns:
        stats['total_deductions'] = monthly_df['총공제액'].sum()
    else:
        stats['total_deductions'] = 0
    
    # 총 실수령액
    if '실수령액' in monthly_df.columns:
        stats['total_net'] = monthly_df['실수령액'].sum()
    else:
        stats['total_net'] = 0
    
    # 총 회사부담금
    if '총회사부담금' in monthly_df.columns:
        stats['total_obligations'] = monthly_df['총회사부담금'].sum()
    else:
        stats['total_obligations'] = 0
    
    # 총 인건비 = 지급액 + 회사부담금 (실제 회사가 부담하는 총 비용)
    stats['total_labor_cost'] = stats['total_earnings'] + stats['total_obligations']
    
    # 인당 평균 비용
    if stats.get('total_fte', 0) > 0:
        stats['avg_cost_per_person'] = stats['total_labor_cost'] / stats['total_fte']
    else:
        stats['avg_cost_per_person'] = 0
    
    # 부서별 요약
    if '부서' in monthly_df.columns:
        dept_agg = {'사원번호': 'count'}
        if 'FTE' in monthly_df.columns:
            dept_agg['FTE'] = 'sum'
        if '총지급액' in monthly_df.columns:
            dept_agg['총지급액'] = 'sum'
        if '총공제액' in monthly_df.columns:
            dept_agg['총공제액'] = 'sum'
        if '실수령액' in monthly_df.columns:
            dept_agg['실수령액'] = 'sum'
        if '총회사부담금' in monthly_df.columns:
            dept_agg['총회사부담금'] = 'sum'
        
        dept_summary = monthly_df.groupby('부서').agg(dept_agg).reset_index()
        dept_summary = dept_summary.rename(columns={'사원번호': '현원'})
        
        if '총지급액' in dept_summary.columns and '총회사부담금' in dept_summary.columns:
            dept_summary['총인건비'] = dept_summary['총지급액'] + dept_summary['총회사부담금']
        elif '총지급액' in dept_summary.columns:
            dept_summary['총인건비'] = dept_summary['총지급액']
        
        if '총인건비' in dept_summary.columns and 'FTE' in dept_summary.columns:
            dept_summary['인당평균'] = dept_summary['총인건비'] / dept_summary['FTE']
        elif '총인건비' in dept_summary.columns and '현원' in dept_summary.columns:
            dept_summary['인당평균'] = dept_summary['총인건비'] / dept_summary['현원']
        
        dept_summary = dept_summary.sort_values('총인건비', ascending=False) if '총인건비' in dept_summary.columns else dept_summary
        stats['dept_summary'] = dept_summary
    
    return stats


# ===================================================================
# 4. 급여 항목 분류 유틸리티
# ===================================================================

def classify_payroll_columns(df: pd.DataFrame) -> Dict[str, List[str]]:
    """
    DataFrame 컬럼을 지급/공제/회사부담/메타 그룹으로 분류합니다.
    
    Returns:
        {
            'earnings': [지급 항목 컬럼 리스트],
            'deductions': [공제 항목 컬럼 리스트],
            'obligations': [회사부담 항목 컬럼 리스트],
            'meta': [메타데이터 컬럼 리스트],
            'summary': [합계 컬럼 리스트],
        }
    """
    # 한국어 번역된 값의 역매핑
    earning_vals = set(EARNING_MAP.values())
    deduction_vals = set(DEDUCTION_MAP.values())
    obligation_vals = set(OBLIGATION_MAP.values())
    
    meta_cols = ['사원번호', '직원명', '부서(원문)', '부서', 'Reg_Pat', '급여유형', '기간번호']
    summary_cols = ['총지급액', '총공제액', '실수령액', '총회사부담금', '기타수당', '기타공제', '기타회사부담']
    
    result = {
        'earnings': [],
        'deductions': [],
        'obligations': [],
        'meta': [],
        'summary': [],
    }
    
    for col in df.columns:
        if col in meta_cols:
            result['meta'].append(col)
        elif col in summary_cols:
            result['summary'].append(col)
        elif col in earning_vals:
            result['earnings'].append(col)
        elif col in deduction_vals:
            result['deductions'].append(col)
        elif col in obligation_vals:
            result['obligations'].append(col)
        # 원문 스페인어 매칭도 시도
        elif col in EARNING_MAP:
            result['earnings'].append(col)
        elif col in DEDUCTION_MAP:
            result['deductions'].append(col)
        elif col in OBLIGATION_MAP:
            result['obligations'].append(col)
    
    return result


# ===================================================================
# 유틸리티: MXN → KRW 환산
# ===================================================================

MXN_TO_KRW_RATE = 80.0  # 2026년 3월 기준 대략적 환율 (1 MXN ≈ 80 KRW)

def mxn_to_krw(amount: float) -> float:
    """멕시코 페소(MXN)를 한국 원(KRW)으로 환산."""
    return amount * MXN_TO_KRW_RATE


def format_mxn(amount: float) -> str:
    """MXN 금액을 포맷팅 (예: $12,345.67)"""
    return f"${amount:,.2f}"


def format_krw(amount: float) -> str:
    """KRW 금액을 포맷팅 (예: ₩987,654)"""
    return f"₩{amount:,.0f}"
